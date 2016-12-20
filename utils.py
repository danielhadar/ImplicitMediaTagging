import numpy as np
import sys
import dictionaries
from scipy.stats import pearsonr
from scipy.stats import spearmanr
from scipy.stats import skew
from scipy.stats import kurtosis
from sklearn.metrics import r2_score
from sklearn.metrics import accuracy_score
from sklearn.metrics import f1_score
from sklearn.metrics import matthews_corrcoef
from math import isnan
from scipy import cluster
from sklearn.decomposition import PCA
from sklearn.cross_validation import LeavePLabelOut
from sklearn import cluster as sklearn_cluster
import pandas as pd

global SUBJECTS_IDS
global PICKLES_FOLDER

def warn(*args, **kwargs):
    pass
import warnings
warnings.warn = warn

fps = 24
NUM_CLIPS = 18

# PARENT_FOLDER = '/Volumes/MyPassport/phase_b/'      # change here where changing machine
PARENT_FOLDER = '/cs/img/danielhadar/'

RATINGS_DIR = PARENT_FOLDER + 'subjects_ratings/'
DATA_FOLDER = PARENT_FOLDER + 'raw_and_rest_data/'
CSV_FOLDER = PARENT_FOLDER + 'csv/'
LOG_FOLDER = PARENT_FOLDER + 'logs/'
OBJECTIVE_FOLDER = '/Users/danielhadar/Documents/Thesis/ExperimentCode/metadata/subject rating PhaseA'

BLENDSHAPES = ['EyeBlink_L', 'EyeBlink_R', 'EyeSquint_L', 'EyeSquint_R', 'EyeDown_L', 'EyeDown_R', 'EyeIn_L', 'EyeIn_R',
               'EyeOpen_L', 'EyeOpen_R', 'EyeOut_L', 'EyeOut_R', 'EyeUp_L', 'EyeUp_R', 'BrowsD_L', 'BrowsD_R',
               'BrowsU_C', 'BrowsU_L', 'BrowsU_R', 'JawOpen', 'LipsTogether', 'JawLeft', 'JawRight', 'JawFwd',
               'LipsUpperUp_L', 'LipsUpperUp_R', 'LipsLowerDown_L', 'LipsLowerDown_R', 'LipsUpperClose',
               'LipsLowerClose', 'MouthSmile_L', 'MouthSmile_R', 'MouthDimple_L', 'MouthDimple_R', 'LipsStretch_L',
               'LipsStretch_R', 'MouthFrown_L', 'MouthFrown_R', 'MouthPress_L', 'MouthPress_R', 'LipsPucker',
               'LipsFunnel', 'MouthLeft', 'MouthRight', 'ChinLowerRaise', 'ChinUpperRaise', 'Sneer_L', 'Sneer_R',
               'Puff', 'CheekSquint_L', 'CheekSquint_R']    # len = 51

GOOD_BLENDSHAPES = ['EyeBlink_L', 'EyeBlink_R','EyeIn_L', 'EyeIn_R', 'BrowsU_C', 'BrowsU_L', 'BrowsU_R', 'JawOpen', 'MouthLeft',
                    'MouthRight', 'MouthFrown_L', 'MouthFrown_R', 'MouthSmile_L', 'MouthSmile_R', 'MouthDimple_L',
                    'MouthDimple_R', 'LipsStretch_L', 'LipsStretch_R', 'LipsUpperUp', 'LipsFunnel', 'ChinLowerRaise',
                    'Sneer', 'CheekSquint_L', 'CheekSquint_R']      # len = 24

MY_BS = ['EyeBlink_L', 'EyeBlink_R', 'MouthSmile_L', 'MouthSmile_R', 'MouthDimple_L', 'MouthDimple_R', 'LipsStretch_L',
               'LipsStretch_R', 'Sneer_L', 'Sneer_R']

SUBJECTS_DICT = {315823492: [6, 10], 315688713: [4, 10], 337835383: [8, 11], 200398733: [6, 15],
                 308286285: [6, 9], 301840336: [4, 14], 336079314: [4, 11], 203667092: [11, 11],
                 304957913: [5, 9], 304854938: [4, 10], 311461917: [5, 10], 203712351: [5, 15],
                 304835366: [5, 10], 332521830: [11, 15], 203237607: [4, 17], 311357735: [5, 14],
                 305584989: [12, 16], 308476639: [5, 14], 204033971: [4, 9], 312282494: [4, 10],
                 203931639: [5, 24], 204713721: [4, 15], 321720443: [6, 15], 317857084: [5, 10],
                 204058010: [5, 10], 203025663: [5, 9]}     # len = 26

MOUTH_BS = ['JawOpen', 'LipsTogether', 'LipsUpperUp_L', 'LipsUpperUp_R', 'LipsLowerDown_L', 'LipsLowerDown_R', 'LipsUpperClose',
               'LipsLowerClose', 'MouthSmile_L', 'MouthSmile_R', 'MouthDimple_L', 'MouthDimple_R', 'LipsStretch_L',
               'LipsStretch_R', 'MouthFrown_L', 'MouthFrown_R', 'MouthPress_L', 'MouthPress_R', 'LipsPucker',
               'LipsFunnel', 'MouthLeft', 'MouthRight', 'ChinLowerRaise', 'ChinUpperRaise']

EYES_AREA_BS = ['EyeBlink_L', 'EyeBlink_R', 'EyeSquint_L', 'EyeSquint_R', 'BrowsD_L', 'BrowsD_R',
               'BrowsU_C', 'BrowsU_L', 'BrowsU_R']

SMILE_BS = ['MouthSmile_L', 'MouthSmile_R']

BLINKS_BS = ['EyeBlink_L', 'EyeBlink_R']

inf = float('Inf')

def flatten_list(l):
    if np.ndim(l) == 1:
        return l
    return [float(item) for sublist in l for item in sublist]


def slice_features_df_for_specific_blendshapes(df, blendshapes_list):

    new_columns_list = []

    for col in df.columns.values:
        if col == 'time':
            new_columns_list.append('time')
        elif col == 'is_locked':
            new_columns_list.append('is_locked')
        elif col == 'ind':
            new_columns_list.append('ind')
        elif col == 'new_response_type':
            new_columns_list.append('new_response_type')

        else:
            for b in blendshapes_list:
                if b in col:
                    new_columns_list.append(col)
                    break

    return df.ix[:,new_columns_list].copy()

def scale(val):
    """
    Scale the given value from the scale of val to the scale of bot-top.
    """
    bot = 0
    top = 1

    if max(val)-min(val) == 0:
        return val
    return ((val - min(val)) / (max(val)-min(val))) * (top-bot) + bot

def scale_list(l):
    """
    Scale the given value from the scale of val to the scale of bot-top.
    """
    bot = 0
    top = 1

    if max(l)-min(l) == 0:
        return l
    for idx,val in enumerate(l):
        l[idx] = ((val - min(l)) / (max(l)-min(l))) * (top-bot) + bot
    return l

def my_pow(val):
    return val**2

def unique_sequences_in_list(l):
    # [2,1,2,3,3,1,1,1,3,3,3,1] -> [2, 1, 2, 3, 1, 3, 1]
    ret_list = [l[0]]

    for i in range(1, len(l)):
        if l[i] == l[i-1]:
            continue
        else:
            ret_list.append(l[i])

    return ret_list

def list_to_unique_list_preserve_order(list):
    # http://stackoverflow.com/questions/480214/how-do-you-remove-duplicates-from-a-list-in-python-whilst-preserving-order
    seen = set()
    seen_add = seen.add
    return [i for i in list if not (i in seen or seen_add(i))]

def grouper(iterable, n, fillvalue=None):
    from itertools import zip_longest
    args = [iter(iterable)] * n
    return zip_longest(*args, fillvalue=fillvalue)


def export_df_to_pickle(df, export_path):
    df.to_pickle(export_path)


def load_pickle_to_df(import_path):
    import pandas as pd
    return pd.read_pickle(import_path)


def load_pickle_list_to_df(import_path, subj_id_and_axis):
    import os
    dfs = []

    for dirname, dirnames, filenames in os.walk(import_path):
        for filename in sorted(filenames):
            if subj_id_and_axis in filename:
                dfs.append(load_pickle_to_df(os.path.join(dirname, filename)))

    return dfs


def export_dict_to_pickle(dict, export_path):
    import pickle
    with open(export_path, 'wb') as handle:
        pickle.dump(dict, handle)


def load_pickle_to_dict(import_path):
    import pickle
    with open(import_path, 'rb') as handle:
        return pickle.load(handle)


def find_peaks(v, delta=0.1, x = None):
    """
    Converted from MATLAB script at http://billauer.co.il/peakdet.html

    Returns two arrays

    function [maxtab, mintab]=peakdet(v, delta, x)
    %PEAKDET Detect peaks in a vector
    %        [MAXTAB, MINTAB] = PEAKDET(V, DELTA) finds the local
    %        maxima and minima ("peaks") in the vector V.
    %        MAXTAB and MINTAB consists of two columns. Column 1
    %        contains indices in V, and column 2 the found values.
    %
    %        With [MAXTAB, MINTAB] = PEAKDET(V, DELTA, X) the indices
    %        in MAXTAB and MINTAB are replaced with the corresponding
    %        X-values.
    %
    %        A point is considered a maximum peak if it has the maximal
    %        value, and was preceded (to the left) by a value lower by
    %        DELTA.

    % Eli Billauer, 3.4.05 (Explicitly not copyrighted).
    % This function is released to the public domain; Any use is allowed.

    """
    maxtab = []
    mintab = []

    if x is None:
        x = np.arange(len(v))

    v = np.asarray(v)

    if len(v) != len(x):
        sys.exit('Input vectors v and x must have same length')

    if not np.isscalar(delta):
        sys.exit('Input argument delta must be a scalar')

    if delta < 0:
        sys.exit('Input argument delta must be positive')

    mn, mx = np.Inf, -np.Inf
    mnpos, mxpos = np.NaN, np.NaN

    lookformax = True

    for i in np.arange(len(v)):
        this = v[i]
        if this > mx:
            mx = this
            mxpos = x[i]
        if this < mn:
            mn = this
            mnpos = x[i]

        if lookformax:
            if this < mx-delta:
            # if this < delta:
                maxtab.append((mxpos, mx))
                mn = this
                mnpos = x[i]
                lookformax = False
        else:
            if this > mn+delta:
            # if this > delta:
                mintab.append((mnpos, mn))
                mx = this
                mxpos = x[i]
                lookformax = True

    # return np.array(maxtab), np.array(mintab)
    return np.array(maxtab)

def count_peaks(v, delta=0.1, x = None):
    return len(find_peaks(v, delta, x))

def get_majoity(list, bin=False, th=2):
    # returns the majority vote.
    # bin=True yields a binary list (2-class: 0/1) - i.e. thresholding
    if bin:
        return 1 if np.mean(list) > th else 0
    else:
        return max(set(list), key=list.count)

def get_pos_or_neg(list):
    # returns a quantized yes/no (for rewatch, likeability) - 0 or 1
    return_list = []
    for i in list:
        if i >= 2:
            return_list.append(1.0)
        else:
            return_list.append(0.0)
    return return_list

def se_of_regression(actual, predicted):
    # Standard error of the regression
    # https://www.otexts.org/fpp/4/4
    N = len(predicted)
    return np.sqrt(
        (1/(N-2)) * sum([pow((a-b),2) for a,b in zip(predicted, actual)])
    )

def balanced_accuracy_score(actual, predicted):
    # balanced accuracy score: https://en.wikipedia.org/wiki/Evaluation_of_binary_classifiers
    # (TP/P + TN/N)/2
    # matrix is: [[TN,FN],[FP,TP]]
    from sklearn.metrics import confusion_matrix
    [[tn, fn],[fp, tp]] = confusion_matrix(predicted, actual)

    return (tp/(tp+fp) + tn/(tn+fn))/2

def previous_and_next_iterator(iterable):
    # iterates over 'iterable' while allowing access to prev and next
    # http://stackoverflow.com/questions/1011938/python-previous-and-next-values-inside-a-loop
    # modified for df.groupby, changed None of beginning and end to tuples of (None,None)
    from itertools import tee, islice, chain

    prevs, items, nexts = tee(iterable, 3)
    prevs = chain([(None, None)], prevs)
    nexts = chain(islice(nexts, 1, None), [(None, None)])
    return zip(prevs, items, nexts)

def slip_by_underline(str):
    return str.split('_')[0]

def add_original_clip(df):
    # adds index column 'org_clip' based upon index column 'clip_id' and return ['subj_id', 'clip_id', 'org_clip']
    df['org_clip'] = df.index.get_level_values('clip_id')
    df['org_clip'] = df['org_clip'].apply(lambda x: int(x.split('_')[0]))
    return df.set_index('org_clip', append=True).reorder_levels(['subj_id', 'clip_id', 'org_clip'])

def add_y(df, y_df, axis):
    # adds y ratings to the df
    df = df.reset_index('clip_id')
    for clip in y_df.index:
        df.loc[clip, axis[0].strip()] = y_df.loc[clip, axis[0].strip()]
    return df.set_index('clip_id', append=True)

def scale_column_by(df, column_name, scale_by, is_majority_vote):
    if scale_by:
        df[column_name] = df.groupby(level=scale_by)[column_name].apply(scale)
    if is_majority_vote:
        df[column_name] = df[column_name].apply(np.round)
    return df


def hl_win_location_table():
    import pandas as pd
    import numpy as np

    df = pd.read_pickle(dictionaries.PICKLES_FOLDER + '/org_raw_with_hl.pickle')
    res = pd.DataFrame(index=pd.MultiIndex(levels=[[], []], labels=[[], []], names=['subj_id', 'org_clip']), columns=['to_end', 'relative_to_start', 'relative_to_end'])

    for subj in dictionaries.SUBJECTS_IDS:
        for clip in dictionaries.CLIPS:
            cur_df = df.loc[(subj,clip)]

            start_watch = float(cur_df.loc['watch'].time.head(1).values)
            # end_watch = float(cur_df.loc['watch'].time.tail(1).values)
            end_watch = start_watch + dictionaries.CLIPS_AND_TIMES[clip]
            start_hl = float(cur_df.loc['hl'].time.head(1).values)
            end_hl = float(cur_df.loc['hl'].time.tail(1).values)

            res.loc[(subj, clip), ['to_end', 'relative_to_start', 'relative_to_end']] = pd.Series([
                end_watch - end_hl - 3,
                (start_hl+3 - start_watch) / (end_watch - start_watch),
                (end_watch - (end_hl+3)) / (end_watch - start_watch)
            ]).values

            # res.loc[(subj, clip), ['from_start', 'to_end', 'relative_to_start', 'relative_to_end']] = pd.Series([
            #     float(cur_df.loc['hl'].time.head(1).values) - float(cur_df.loc['watch'].time.head(1).values),
            #     float(cur_df.loc['watch'].time.tail(1).values) - float(cur_df.loc['hl'].time.tail(1).values),
            #
            #     float((float(cur_df.loc['hl'].time.head(1).values)+3 - float(cur_df.loc['watch'].time.head(1).values)) / (float(cur_df.loc['watch'].time.tail(1).values) - float(cur_df.loc['watch'].time.head(1).values))),
            #     float((float(cur_df.loc['watch'].time.tail(1).values) - float(cur_df.loc['hl'].time.tail(1).values)-3) / (float(cur_df.loc['watch'].time.tail(1).values) - float(cur_df.loc['watch'].time.head(1).values)))
            # ]).values

    res.to_csv('hl_win_location.csv')

# if __name__ == '__main__':
#     # hl_win_location_table()
#     import learning
#     import re
#     id_pattern = re.compile("\d{9}")
#
#     # for name in ['cv_results_df_valence_6.csv', 'cv_results_df_arousal_7.csv', 'cv_results_df_likeability_8.csv', 'cv_results_df_rewatch_9.csv']:
#     for name in ['cv_results_df_valence_6.csv']:
#     # for name in ['cv_results_df_arousal_7.csv']:
#         f = open(LOG_FOLDER + name)
#         all_predicted_y = []
#         all_actual_y = []
#
#         for line in f:
#             line = line.split(',')
#
#             if line[0] == '200398733':
#                 arr = [float(line[2])]
#             elif id_pattern.match(line[0]):
#                 arr.append(float(line[2]))
#
#             if line[0] == '337835383':
#                 # all_predicted_y.append([np.mean(arr), np.median(arr)])
#                 all_predicted_y.append(np.mean(arr))
#                 all_actual_y.append(float(line[3]))
#
#         print(name, pearsonr(all_actual_y, all_predicted_y))
#
#         pred = []
#         act = []
#         for idx in range(len(all_actual_y)):
#             clf = learning.run_learning(all_predicted_y[:idx] + all_predicted_y[idx+1:], all_actual_y[:idx] + all_actual_y[idx+1:], 'linear_regression')
#             pred.append(clf.predict(all_predicted_y[idx])[0])
#             act.append(all_actual_y[idx])
#         print(name, pearsonr(pred,act))

def quantize_list(l, th, dist, env):
    return_list = np.zeros(len(l))

    if env:         # around *dist* from th gets '-1'
        for idx, num in enumerate(l):
            if num < (th-dist):
                return_list[idx] = 0
            elif num > (th+dist):
                return_list[idx] = 1
            else:
                return_list[idx] = -1

    else:
        for idx, num in enumerate(l):
            if num < th:
                return_list[idx] = 0
            else:
                return_list[idx] = 1

    return return_list


def calc_binary_from_cv_output(path, filename, discard_middle, leave_subj_out=True):
    import openpyxl

    wb = openpyxl.load_workbook(path + filename)
    # averages = {'V':np.zeros(3), 'A':np.zeros(3), 'L':np.zeros(3), 'R':np.zeros(3)}

    for sheet_name in ['V', 'A', 'L', 'R']:
        outliers = 0
        print(sheet_name)
        sheet = wb.get_sheet_by_name(sheet_name)
        _ = sheet.cell(row=521 if leave_subj_out else 505, column=1, value='end')  # so iter_rows gets to final subject's last line

        for idx, row in enumerate(sheet.iter_rows()):
            if row[0].value == 'end':
                break
            elif row[0].value and not row[1].value:   # subject (w.l.o.g clip) identifier row: "10 50 True 0.695545117716 0.811028369751 311461917 (36)"
                id = row[0].value.split(' ')[-1]
                print(id)
                chunk_start_idx = idx + 2
                cur_actual = []
                cur_predicted = []
            elif row[0].value and row[1].value:     # within subj/clip
                cur_actual.append(row[2].value)
                cur_predicted.append(row[3].value)
            else:                                   # new line <br> between subjects/clips
                quantized_cur_actual = quantize_list(cur_actual, np.mean(cur_actual), np.std(cur_actual)/2, discard_middle)
                quantized_cur_predicted = quantize_list(cur_predicted, np.mean(cur_predicted), np.std(cur_predicted)/2, False)
                tp = 0
                tn = 0
                fp = 0
                fn = 0

                for i,j in enumerate(range(chunk_start_idx, idx+1)):
                    y_actual = quantized_cur_actual[i]
                    y_predicted = quantized_cur_predicted[i]
                    _ = sheet.cell(row=j, column=6, value=y_actual)
                    _ = sheet.cell(row=j, column=7, value=y_predicted)
                    if y_predicted == -1 or y_actual == -1:
                        outliers += 1
                        continue
                    elif y_actual == y_predicted == 1:
                        tp += 1
                    elif y_actual == y_predicted == 0:
                        tn += 1
                    elif y_actual > y_predicted:
                        fp += 1
                    else:
                        fn += 1

                print(tp, tn, fp, fn)
                acc = (tp+tn)/(tp+tn+fp+fn)
                # bacc = ((tp/(tp+fp))+(tn/(tn+fn)))/2
                mcc = (tp*tn-fp*fn)/np.sqrt((tp+fp)*(tp+fn)*(tn+fp)*(tn+fn))
                _ = sheet.cell(row=j, column=8, value=acc)
                # _ = sheet.cell(row=j, column=9, value=bacc)
                _ = sheet.cell(row=j, column=10, value=mcc)

        #         averages[sheet_name][0].append(acc)
        #         averages[sheet_name][1].append(bacc)
        #         averages[sheet_name][2].append(mcc)
        #
        # print(sheet_name)
        # print(np.mean([i[0] for i in averages]))
        # print(np.mean([i[1] for i in averages]))
        # print(np.mean([i[2] for i in averages]))

    wb.save(filename=path + 'binarization_after.xlsx')
    print(outliers)




if __name__ == '__main__':
    # calc_binary_from_cv_output(LOG_FOLDER + '/obj_rank_leave_clip_out/notmodel4each/', 'for_binarization.xlsx', discard_middle=True, leave_subj_out=False)

    import re
    id_pattern = re.compile("\d{9}")

    # for name in ['cv_results_df_valence_6.csv', 'cv_results_df_arousal_5.csv', 'cv_results_df_likeability_8.csv', 'cv_results_df_rewatch_9.csv']:       # obj
    # for name in ['cv_results_df_valence_.csv', 'cv_results_df_arousal_2.csv', 'cv_results_df_likeability_3.csv', 'cv_results_df_rewatch_4.csv']:      # subj
    for name in ['temp.csv']:      # subj
        f = open(LOG_FOLDER + 'obj_rank_leave_clip_out/notmodel4each/' + name)
        # f = open(LOG_FOLDER + 'subj_rank_leave_clip_out/normodel4each/' + name)
        predicted = {}
        actual = {}
        all_predicted_y = []
        all_actual_y = []


        # average of correlations per subject (suitable for both obj and subj)
        # for line in f:
        #     line = line.split(',')
        #
        #     if line[0] in dictionaries.SUBJECTS_IDS:
        #         try:
        #             predicted[line[0]].append(float(line[2]))
        #             actual[line[0]].append(float(line[3]))
        #         except KeyError:
        #             predicted[line[0]] = [float(line[2])]
        #             actual[line[0]] = [float(line[3])]
        #
        # cur = []
        # for key,val in predicted.items():
        #     cur.append(pearsonr(val, actual[key])[0])
        #
        # print(name, np.mean(cur), np.std(cur))


        # average predicted per clip and calculate correlations over all clips (suitable just for obj)
        for line in f:
            line = line.split(',')

            if line[0] == '200398733':
                if float(line[2]) != -1:
                    arr = [float(line[2])]
                else:
                    arr = []
            elif id_pattern.match(line[0]):
                if float(line[2]) != -1:
                    arr.append(float(line[2]))

            if line[0] == '337835383':
                # all_predicted_y.append(np.mean(arr))        # for pearson r
                all_predicted_y.append(get_majoity(arr))      # for binary accuracy
                all_actual_y.append(float(line[3]))

        # for pearson r
        print(np.std(all_predicted_y))
        print(name, pearsonr(all_actual_y, all_predicted_y))

        # for binary prediction
        tp = 0
        tn = 0
        fp = 0
        fn = 0

        for idx in range(len(all_predicted_y)):
            y_actual = all_actual_y[idx]
            y_predicted = all_predicted_y[idx]
            if y_actual == y_predicted == 1:
                tp += 1
            elif y_actual == y_predicted == 0:
                tn += 1
            elif y_actual > y_predicted:
                fp += 1
            else:
                fn += 1

        acc = (tp+tn)/(tp+tn+fp+fn)
        print(acc)